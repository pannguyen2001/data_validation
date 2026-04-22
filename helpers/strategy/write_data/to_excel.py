import os
import openpyxl
import pandas as pd
from copy import copy
from loguru import logger
from utils.logger_wrapper import logger_wrapper
from string import Template
from .base_strategy import WriteDataStrategy


write_successfully_template = Template("""Write data to excel file successfully.
Detail info:
    Sheet name: ${sheet_name}.
    File: ${file_path}.
    Mode: ${mode_flag}.
    If sheet exists: ${if_sheet_exists}.""")

class WriteToExcelStrategy(WriteDataStrategy):
    # def __init__(self, file_path: str, sheet_name: str):
    #     super().__init__(file_path, sheet_name)

    @logger_wrapper
    def write_data(
        self,
        df: pd.DataFrame,
        file_path: str,
        sheet_name: str = "Sheet1",
        origin_file_path: str = "",
        index: bool = False,
        mode: str = "replace",
        *args,
        **kwargs
        ) -> None:
            logger.info(f"[{self.__class__.__name__}] Write data to excel file.")
            mode_flag = "w" if not os.path.exists(file_path) else "a"
            if_sheet_exists = "replace" if mode == "replace" else "overlay"
            config = {
                "engine": kwargs.get("engine") or "openpyxl",
                "mode": mode_flag,
                **kwargs
            }
            if mode_flag == "a":
                config["if_sheet_exists"] = if_sheet_exists
            with pd.ExcelWriter(file_path, **config) as writer:
                df.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=index,
                    *args,
                    **kwargs
                    )

            if origin_file_path:
                format_excel_file(src_file=origin_file_path, dest_file=file_path, sheet_name=sheet_name)

            logger.success(write_successfully_template.safe_substitute(
                sheet_name=sheet_name,
                file_path=file_path,
                mode_flag=mode_flag,
                if_sheet_exists=if_sheet_exists
            ))

@logger_wrapper
def format_excel_file(src_file: str = "", dest_file: str = "", sheet_name: str= "") -> None:
    # Load your workbook
    src_wb = openpyxl.load_workbook(src_file)
    src_ws = src_wb[sheet_name]
    dest_wb = openpyxl.load_workbook(dest_file)
    dest_ws = dest_wb[sheet_name]

    # Define the header row (typically row 1)
    header_row_index = 1

    # Iterate through cells in the header row
    for col in range(1, src_ws.max_column + 1):
        src_cell = src_ws.cell(row=header_row_index, column=col)
        dest_cell = dest_ws.cell(row=header_row_index, column=col)
        
        # Copy value (optional)
        dest_cell.value = src_cell.value
        
        # Copy all style attributes if they exist
        if src_cell.has_style:
            dest_cell.font = copy(src_cell.font)
            dest_cell.border = copy(src_cell.border)
            dest_cell.fill = copy(src_cell.fill)
            dest_cell.number_format = copy(src_cell.number_format)
            dest_cell.protection = copy(src_cell.protection)
            dest_cell.alignment = copy(src_cell.alignment)

    # Save the workbook
    dest_wb.save(dest_file)

'''
# Write excel with format
import os
import pandas as pd
from loguru import logger
from string import Template
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from utils.logger_wrapper import logger_wrapper
from .base_strategy import WriteDataStrategy


write_successfully_template = Template("""Write data to excel file successfully.
Detail info:
    Sheet name: ${sheet_name}.
    File: ${file_path}.
    Mode: ${mode_flag}.
    If sheet exists: ${if_sheet_exists}.""")


class WriteToExcelStrategy(WriteDataStrategy):

    @logger_wrapper
    def write_data(
        self,
        df: pd.DataFrame,
        file_path: str,
        sheet_name: str = "Sheet1",
        index: bool = False,
        mode: str = "replace",
        *args,
        **kwargs
    ) -> None:
        logger.info(f"[{self.__class__.__name__}] Write data to excel file.")

        mode_flag = "w" if not os.path.exists(file_path) else "a"
        if_sheet_exists = "replace" if mode == "replace" else "overlay"

        # Separate writer kwargs from to_excel kwargs
        writer_engine = kwargs.pop("engine", "openpyxl")

        writer_kwargs = {
            "engine": writer_engine,
            "mode": mode_flag,
        }

        if mode_flag == "a":
            writer_kwargs["if_sheet_exists"] = if_sheet_exists

        with pd.ExcelWriter(file_path, **writer_kwargs) as writer:
            df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=index,
                *args,
                **kwargs
            )

        # Apply Excel-native formatting
        self._format_worksheet(
            file_path=file_path,
            sheet_name=sheet_name,
            df=df,
            index=index
        )

        logger.success(write_successfully_template.safe_substitute(
            sheet_name=sheet_name,
            file_path=file_path,
            mode_flag=mode_flag,
            if_sheet_exists=if_sheet_exists
        ))

    def _format_worksheet(
        self,
        file_path: str,
        sheet_name: str,
        df: pd.DataFrame,
        index: bool = False
    ) -> None:
        wb = load_workbook(file_path)
        ws = wb[sheet_name]

        # 1) Freeze first row + first column
        # This means row 1 and column A stay visible
        ws.freeze_panes = "B2"

        # 2) Dark blue header row
        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")  # dark blue
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # 3) Optional: auto width for all columns
        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter

            for cell in column_cells:
                try:
                    cell_value = "" if cell.value is None else str(cell.value)
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except Exception:
                    pass

            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        wb.save(file_path)
'''