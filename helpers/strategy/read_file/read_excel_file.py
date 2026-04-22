import polars as pl
from typing import Union, List
from .base_strategy import ReadDataStrategy
from utils.logger import logger
from openpyxl import load_workbook
from python_calamine import CalamineWorkbook


def load_column_names(file_path: str, sheet_name: str) -> List[str]:
    workbook = CalamineWorkbook.from_path(file_path)
    sheet = workbook.get_sheet_by_name(sheet_name)
    sheet_data = sheet.to_python()
    columns = [col for col in sheet_data[0] if col is not None and col]
    return columns

# def load_column_names(file_path: str, sheet_name: str) -> List[str]:
#     wb = load_workbook(file_path, read_only=True)
#     ws = wb[sheet_name]
#     columns = [cell.value for cell in next(ws.iter_rows(max_row=1))]
#     return columns


class ReadExcelFileStrategy(ReadDataStrategy):
    def __init__(self, file_path: str = "") -> None:
        super().__init__(file_path)

    def load(self, sheet_name: Union[str, List[str]], *args, **kwargs):
        usecols = kwargs.get("usecols")
        # Read 0 rows to get column names
        # columns = pd.ExcelFile(self.file_path).parse(sheet_name, nrows=0)
        # columns: List = load_column_names(self.file_path, sheet_name)
        # schema = {col: pl.Utf8 for col in columns if col is not None and col}
        # if usecols is not None:
        #     schema = {col: pl.Utf8 for col in usecols if col in columns}

        df = pl.read_excel(
            self.file_path,
            sheet_name=sheet_name,
            has_header=True,
            # schema_overrides=schema,
            columns=usecols,
            infer_schema_length=0
        )
        df = df.to_pandas()

        # if usecols is not None:
        #     invalid_columns = set(usecols) - set(df.columns)
        #     if invalid_columns:
        #         logger.error(f"[{self.__class__.__name__} - {sheet_name}] Column name not in data columns: {', '.join(invalid_columns)}. Data columns: {df.columns.value.tolist()}")
        #     df = df[usecols]

        logger.success(
            f"[{self.__class__.__name__} - {sheet_name}] Data loaded successfully: {df.shape[0]:,} rows, {df.shape[1]:,} columns."
        )
        return df

    # def load(self, sheet_name: Union[str, List[str]], *args, **kwargs):
    #     chunk_size = kwargs.get("chunk_size", 10_000)
    #     usecols = kwargs.get("usecols")
    #     # Read the header separately to maintain column consistency
    #     with pd.ExcelFile(self.file_path, engine="calamine") as xf:
    #         columns = xf.parse(sheet_name, nrows=0).columns.tolist()
    #         if usecols is not None:
    #             invalid_columns = [col for col in usecols if col not in columns]
    #             if invalid_columns:
    #                 logger.error(
    #                     f"[{self.__class__.__name__}] Invalid columns: {invalid_columns}"
    #                 )
    #                 return

    #         skiprows = 1
    #         chunks = []
    #         while True:
    #             df_chunk = pd.read_excel(
    #                 xf,
    #                 sheet_name=sheet_name,
    #                 nrows=chunk_size,
    #                 skiprows=skiprows,
    #                 header=None,
    #                 engine="calamine",
    #                 dtype=str,
    #             )
    #             if df_chunk.empty:
    #                 break
    #             df_chunk.columns = columns
    #             if usecols is not None:
    #                 df_chunk = df_chunk[usecols]
    #             chunks.append(df_chunk)
    #             skiprows += chunk_size

    #         if not chunks:
    #             return pd.DataFrame(columns=columns)
    #         df = pd.concat(chunks, ignore_index=True)
    #         logger.success(f"[{self.__class__.__name__} - {sheet_name}] Data loaded successfully: {df.shape[0]:,} rows, {df.shape[1]:,} columns.")
    #         return df
